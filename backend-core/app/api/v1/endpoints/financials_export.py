"""
Financial Statements Excel Export Endpoint
Generates professional Excel files for financial data download.
"""

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from datetime import datetime
import io

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.db.session import db

router = APIRouter()

# Current year for validation
CURRENT_YEAR = datetime.now().year


# Professional cell styles
def setup_styles(wb):
    """Set up professional styles for the workbook."""
    # Header style
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, size=12, color="FFFFFF")
    header_style.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Subtotal style
    subtotal_style = NamedStyle(name="subtotal_style")
    subtotal_style.font = Font(bold=True, size=11)
    subtotal_style.fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    subtotal_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Normal style
    normal_style = NamedStyle(name="normal_style")
    normal_style.font = Font(size=10)
    normal_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Percent style
    percent_style = NamedStyle(name="percent_style")
    percent_style.font = Font(size=10)
    percent_style.number_format = '0.00%'
    percent_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Currency/number style
    currency_style = NamedStyle(name="currency_style")
    currency_style.font = Font(size=10)
    currency_style.number_format = '#,##0.00'
    currency_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Info style (for subtitle)
    info_style = NamedStyle(name="info_style")
    info_style.font = Font(size=10, italic=True, color="666666")
    info_style.alignment = Alignment(horizontal="left")
    
    try:
        wb.add_named_style(header_style)
        wb.add_named_style(subtotal_style)
        wb.add_named_style(normal_style)
        wb.add_named_style(percent_style)
        wb.add_named_style(currency_style)
        wb.add_named_style(info_style)
    except:
        pass  # Styles may already exist
    
    return wb


def format_number(value, is_percent=False):
    """Format a number for Excel."""
    if value is None:
        return ""
    try:
        val = float(value)
        if is_percent:
            # Handle both decimal (0.31) and percentage (31.0) formats
            if abs(val) <= 1:
                return val  # Already in decimal, Excel will format as %
            return val / 100  # Convert to decimal for Excel percent format
        return val
    except:
        return value


def create_sheet(wb, sheet_name, data_rows, years, currency='EGP', fiscal_note=''):
    """Create a formatted sheet with financial data."""
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet names max 31 chars
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(years) + 1)
    title_cell = ws.cell(row=1, column=1, value=f"{sheet_name} (Currency: {currency})")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Subtitle row - "Financials in millions. Fiscal year is January - December."
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(years) + 1)
    subtitle_cell = ws.cell(row=2, column=1, value=fiscal_note)
    subtitle_cell.font = Font(size=10, italic=True, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="center")
    
    # Header row (moved to row 4)
    ws.cell(row=4, column=1, value="Line Item").style = "header_style"
    for i, year in enumerate(years):
        ws.cell(row=4, column=i + 2, value=year).style = "header_style"
    
    # Data rows (start from row 5)
    current_row = 5
    for row_data in data_rows:
        label = row_data.get('label', '')
        values = row_data.get('values', {})
        is_subtotal = row_data.get('isSubtotal', False)
        is_percent = row_data.get('isPercent', False)
        indent = row_data.get('indent', 0)
        
        # Apply indent to label
        display_label = ("    " * indent) + label
        
        # Label cell
        label_cell = ws.cell(row=current_row, column=1, value=display_label)
        if is_subtotal:
            label_cell.style = "subtotal_style"
        else:
            label_cell.style = "normal_style"
        
        # Value cells
        for i, year in enumerate(years):
            val = values.get(year)
            formatted_val = format_number(val, is_percent)
            cell = ws.cell(row=current_row, column=i + 2, value=formatted_val)
            
            if is_subtotal:
                cell.style = "subtotal_style"
            elif is_percent:
                cell.style = "percent_style"
            elif isinstance(formatted_val, (int, float)):
                cell.style = "currency_style"
            else:
                cell.style = "normal_style"
        
        current_row += 1
    
    # Auto-fit column widths
    ws.column_dimensions['A'].width = 45  # Label column
    for i in range(len(years)):
        col_letter = get_column_letter(i + 2)
        ws.column_dimensions[col_letter].width = 15
    
    return ws


@router.get("/financials/{symbol}/export")
async def export_financials(
    symbol: str,
    period_type: str = Query("annual", enum=["annual", "quarterly"]),
    limit: int = Query(10, ge=1, le=20)
):
    """
    Export financial statements to Excel format.
    
    Returns a professionally formatted .xlsx file containing:
    - Income Statement
    - Balance Sheet
    - Cash Flow Statement
    - Financial Ratios
    
    Args:
        symbol: Stock ticker symbol (e.g., COMI, MASR)
        period_type: 'annual' or 'quarterly'
        limit: Number of periods to include (default 10)
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. openpyxl library is not installed."
        )
    
    symbol = symbol.upper().strip()
    
    # Get company info using db singleton
    ticker = await db.fetch_one("""
        SELECT name_en, name_ar, market_code, currency 
        FROM market_tickers WHERE symbol = $1
    """, symbol)
    
    if not ticker:
        raise HTTPException(status_code=404, detail=f"Symbol {symbol} not found")
    
    currency = ticker.get('currency') or 'EGP'
    company_name = ticker.get('name_en') or symbol
    
    # Fetch financial data - FILTER OUT FUTURE YEARS (> current year)
    if period_type == 'annual':
        income_rows = await db.fetch_all(
            f"SELECT * FROM income_statements WHERE symbol = $1 AND period_type = 'annual' AND fiscal_year <= {CURRENT_YEAR} ORDER BY fiscal_year DESC LIMIT {limit}",
            symbol
        )
        balance_rows = await db.fetch_all(
            f"SELECT * FROM balance_sheets WHERE symbol = $1 AND period_type = 'annual' AND fiscal_year <= {CURRENT_YEAR} ORDER BY fiscal_year DESC LIMIT {limit}",
            symbol
        )
        cashflow_rows = await db.fetch_all(
            f"SELECT * FROM cashflow_statements WHERE symbol = $1 AND period_type = 'annual' AND fiscal_year <= {CURRENT_YEAR} ORDER BY fiscal_year DESC LIMIT {limit}",
            symbol
        )
        ratios_rows = await db.fetch_all(
            f"SELECT * FROM financial_ratios_history WHERE symbol = $1 AND fiscal_year <= {CURRENT_YEAR} ORDER BY fiscal_year DESC LIMIT {limit}",
            symbol
        )
    else:
        income_rows = await db.fetch_all(
            f"SELECT * FROM income_statements WHERE symbol = $1 AND period_type = 'quarterly' AND fiscal_year <= {CURRENT_YEAR} ORDER BY fiscal_year DESC, fiscal_quarter DESC LIMIT $2",
            symbol, limit * 4
        )
        balance_rows = await db.fetch_all(
            f"SELECT * FROM balance_sheets WHERE symbol = $1 AND period_type = 'quarterly' AND fiscal_year <= {CURRENT_YEAR} ORDER BY fiscal_year DESC, fiscal_quarter DESC LIMIT $2",
            symbol, limit * 4
        )
        cashflow_rows = await db.fetch_all(
            f"SELECT * FROM cashflow_statements WHERE symbol = $1 AND period_type = 'quarterly' AND fiscal_year <= {CURRENT_YEAR} ORDER BY fiscal_year DESC, fiscal_quarter DESC LIMIT $2",
            symbol, limit * 4
        )
        ratios_rows = []  # Ratios typically annual only
    
    # Import display mappings
    from app.chat.handlers.financials_handler import (
        INCOME_DISPLAY_ORDERED,
        BALANCE_DISPLAY_ORDERED,
        CASHFLOW_DISPLAY_ORDERED,
        RATIOS_DISPLAY_ORDERED
    )
    
    # Process data into export format - INCLUDE ALL ITEMS
    def process_for_export(rows, ordered_list, is_quarterly=False):
        """Convert DB rows to export format. Include ALL items from ordered list."""
        if not rows:
            return [], []
        
        # Build period labels - FILTER OUT INVALID YEARS
        periods = []
        data_by_period = {}
        seen = set()
        
        for r in rows:
            y = r.get('fiscal_year')
            if y is None:
                continue
            y = int(y)
            
            # STRICT year validation: only include 2016 to current year
            # This filters out invalid data like 2076, 2027, 2015
            if y > CURRENT_YEAR or y < 2016:
                continue
                
            y_str = str(y)
            q = r.get('fiscal_quarter')
            label = f"Q{q} {y_str}" if q and is_quarterly else y_str
            if label not in seen:
                seen.add(label)
                periods.append(label)
                data_by_period[label] = dict(r)
        
        # Build processed rows - INCLUDE ALL ITEMS FROM ORDERED LIST
        processed = []
        for col, display_label, options in ordered_list:
            if col == 'period_ending':
                continue
            
            row_obj = {
                'label': display_label,
                'values': {},
                'isSubtotal': options.get('isSubtotal', False),
                'isPercent': options.get('isPercent', False),
                'indent': options.get('indent', 0)
            }
            
            has_any_data = False
            for period in periods:
                val = data_by_period.get(period, {}).get(col)
                if val is not None:
                    try:
                        row_obj['values'][period] = float(val)
                        has_any_data = True
                    except:
                        row_obj['values'][period] = None
                else:
                    row_obj['values'][period] = None
            
            # Smart filtering: Only show rows that have at least ONE value
            # This ensures banking fields don't show for corporate companies and vice versa
            # Matches StockAnalysis.com behavior exactly
            if has_any_data:
                processed.append(row_obj)
        
        return processed, periods
    
    is_quarterly = period_type == 'quarterly'
    income_data, years = process_for_export(income_rows, INCOME_DISPLAY_ORDERED, is_quarterly)
    balance_data, balance_years = process_for_export(balance_rows, BALANCE_DISPLAY_ORDERED, is_quarterly)
    cashflow_data, cf_years = process_for_export(cashflow_rows, CASHFLOW_DISPLAY_ORDERED, is_quarterly)
    ratios_data, _ = process_for_export(ratios_rows, RATIOS_DISPLAY_ORDERED, is_quarterly)
    
    # Use the best years from any available data
    if not years and balance_years:
        years = balance_years
    if not years and cf_years:
        years = cf_years
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    setup_styles(wb)
    
    # Fiscal note for all sheets
    fiscal_note = "Financials in millions. Fiscal year is January - December."
    
    # Create sheets with ALL items
    if income_data and years:
        create_sheet(wb, "Income Statement", income_data, years, currency, fiscal_note)
    if balance_data and years:
        create_sheet(wb, "Balance Sheet", balance_data, years, currency, fiscal_note)
    if cashflow_data and years:
        create_sheet(wb, "Cash Flow Statement", cashflow_data, years, currency, fiscal_note)
    if ratios_data and years:
        create_sheet(wb, "Financial Ratios", ratios_data, years, currency, "Key financial ratios and metrics.")
    
    # If no data, create empty sheet
    if not any([income_data, balance_data, cashflow_data, ratios_data]) or not years:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value=f"No financial data available for {symbol}")
    
    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"{symbol}_Financials_{period_type}_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
